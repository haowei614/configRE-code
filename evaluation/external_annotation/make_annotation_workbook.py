#!/usr/bin/env python3
"""Generate the blinded annotation workbook for the external cases.

Creates Annotation_Sheet.xlsx with:
  - a README tab (annotator metadata + instructions)
  - one tab per case (EHR, SmartGrid, LoanApproval), each listing the 15 agents
    with a Y/N dropdown and a rationale column.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

AGENTS = [
    ("Safety",          "ISO 25010", "Hazard prevention, risk mitigation"),
    ("Performance",     "ISO 25010", "Response time, throughput, capacity"),
    ("Efficiency",      "ISO 25010", "Resource utilization, energy usage"),
    ("Reliability",     "ISO 25010", "Fault tolerance, recoverability"),
    ("Usability",       "ISO 25010", "Learnability, error protection"),
    ("Security",        "ISO 25010", "Authentication, access control"),
    ("Trustworthiness", "ISO 25010", "Privacy, data protection, trust"),
    ("Maintainability", "ISO 25010", "Modularity, testability"),
    ("Compatibility",   "ISO 25010", "Interoperability, co-existence"),
    ("Flexibility",     "ISO 25010", "Adaptability, replaceability"),
    ("Func. Safety",    "ISO 26262", "ASIL levels, hazard analysis"),
    ("Explainability",  "EU AI Act", "Model transparency, interpretability"),
    ("Privacy",         "GDPR",      "Consent, data subject rights"),
    ("Green",           "ISO 14001", "Energy efficiency, carbon footprint"),
    ("Responsibility",  "IEEE 7000", "Ethical accountability, compliance"),
]

CASES = ["EHR", "SmartGrid", "LoanApproval"]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_case_sheet(wb, case):
    ws = wb.create_sheet(title=case)
    headers = ["#", "Agent", "Standard", "Sub-focus", "Relevant (Y/N)", "Rationale (one line)"]
    ws.append(headers)
    style_header(ws, len(headers))

    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = "Please enter Y or N"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "Is a dedicated agent for this concern warranted for THIS system?"
    dv.promptTitle = "Relevant?"
    ws.add_data_validation(dv)

    for i, (name, std, focus) in enumerate(AGENTS, start=1):
        ws.append([i, name, std, focus, "", ""])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)
        dv.add(ws.cell(row=r, column=5))

    widths = [4, 16, 11, 34, 14, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = "A2"
    return ws


def build_readme(wb):
    ws = wb.create_sheet(title="README", index=0)
    lines = [
        ("Blinded Ground-Truth Annotation — External Cases", True),
        ("", False),
        ("Please read 00_INSTRUCTIONS.md and AGENT_POOL.md before starting.", False),
        ("For each case tab, mark every agent Y (relevant) or N (not relevant)", False),
        ("and give a one-line rationale. No fixed count; typical systems need 5-7.", False),
        ("Use ONLY the project descriptions in cases/ and the agent definitions.", False),
        ("Do NOT consult the authors' selections or any tool output.", False),
        ("", False),
        ("Fill in below, then return this file:", True),
    ]
    for text, bold in lines:
        ws.append([text])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=bold, size=13 if bold and ws.max_row == 1 else 11)

    ws.append([])
    for label in ["Annotator ID:", "Background (role / field):", "Years of SE/RE experience:", "Date completed:"]:
        ws.append([label, ""])
        r = ws.max_row
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=r, column=2).border = BORDER
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    for case in CASES:
        build_case_sheet(wb, case)
    out = "Annotation_Sheet.xlsx"
    wb.save(out)
    print(f"Wrote {out} with tabs: README, {', '.join(CASES)}")


if __name__ == "__main__":
    main()
