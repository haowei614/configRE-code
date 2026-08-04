#!/usr/bin/env python3
"""Inter-annotator agreement for the external cases (supports >=2 annotators).

Usage:
  python3 compute_agreement.py Sheet_A1.xlsx Sheet_A2.xlsx Sheet_A3.xlsx

Reports, per case and pooled over all external cases:
  - observed (percent) agreement across annotators
  - Fleiss' kappa            (chance-corrected, multi-rater, nominal Y/N)
  - Krippendorff's alpha     (nominal; robust to raters/missing data)

Optional: if author_ground_truth.json exists and is filled with Y/N, also
reports each annotator's Cohen's kappa vs the author GT and the majority-vote
vs author agreement. Run with no args to (re)generate the GT template.
"""
import itertools
import json
import os
import sys

from openpyxl import load_workbook

AGENTS = [
    "Safety", "Performance", "Efficiency", "Reliability", "Usability",
    "Security", "Trustworthiness", "Maintainability", "Compatibility",
    "Flexibility", "Func. Safety", "Explainability", "Privacy", "Green",
    "Responsibility",
]
CASES = ["EHR", "SmartGrid", "LoanApproval"]
GT_FILE = "author_ground_truth.json"


def init_gt_template():
    tmpl = {case: {a: "" for a in AGENTS} for case in CASES}
    with open(GT_FILE, "w") as f:
        json.dump(tmpl, f, indent=2)
    print(f"Wrote template {GT_FILE}. Optionally fill each entry with 'Y'/'N' "
          f"(author GT) to also get annotator-vs-author agreement.")


def read_annotations(path):
    wb = load_workbook(path, data_only=True)
    out = {}
    for case in CASES:
        ws = wb[case]
        labels = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            agent = str(row[1]).strip()
            val = str(row[4]).strip().upper() if row[4] is not None else ""
            labels[agent] = val if val in ("Y", "N") else ""
        out[case] = labels
    return out


def cohen_kappa(a, b):
    n = len(a)
    if n == 0:
        return float("nan")
    po = sum(1 for x, y in zip(a, b) if x == y) / n
    pe = sum((sum(1 for x in a if x == c) / n) * (sum(1 for y in b if y == c) / n)
             for c in ("Y", "N"))
    return (po - pe) / (1 - pe) if (1 - pe) else float("nan")


def fleiss_kappa(rows):
    """rows: list of items; each item is a list of category labels (one per rater).
    Assumes a fixed number of raters per item."""
    cats = ["Y", "N"]
    n_items = len(rows)
    n_raters = len(rows[0])
    counts = []
    for r in rows:
        counts.append([r.count(c) for c in cats])
    p_j = [sum(counts[i][j] for i in range(n_items)) / (n_items * n_raters)
           for j in range(len(cats))]
    P_i = [(sum(c ** 2 for c in counts[i]) - n_raters) / (n_raters * (n_raters - 1))
           for i in range(n_items)]
    P_bar = sum(P_i) / n_items
    P_e = sum(p ** 2 for p in p_j)
    return (P_bar - P_e) / (1 - P_e) if (1 - P_e) else float("nan")


def krippendorff_alpha_nominal(units):
    """units: list of lists of labels (variable rater count allowed, missing dropped)."""
    pairs = []
    for u in units:
        vals = [v for v in u if v in ("Y", "N")]
        m = len(vals)
        if m < 2:
            continue
        for a, b in itertools.permutations(vals, 2):
            pairs.append((a, b))
    if not pairs:
        return float("nan")
    Do = sum(1 for a, b in pairs if a != b) / len(pairs)
    flat = [v for u in units for v in u if v in ("Y", "N")]
    n = len(flat)
    pe = sum((flat.count(c) * (flat.count(c) - 1)) for c in ("Y", "N")) / (n * (n - 1))
    De = 1 - pe
    return 1 - Do / De if De else float("nan")


def observed_agreement(rows):
    """Average pairwise percent agreement across raters, per item, averaged."""
    tot = 0.0
    for r in rows:
        pairs = list(itertools.combinations(r, 2))
        tot += sum(1 for a, b in pairs if a == b) / len(pairs)
    return tot / len(rows)


def main():
    files = sys.argv[1:]
    if not files:
        print(__doc__)
        if not os.path.exists(GT_FILE):
            init_gt_template()
        return

    anns = [read_annotations(f) for f in files]
    names = [os.path.basename(f) for f in files]
    print(f"Loaded {len(anns)} annotators: {', '.join(names)}")

    gt = None
    if os.path.exists(GT_FILE):
        g = json.load(open(GT_FILE))
        if any(str(g.get(c, {}).get(a, "")).strip().upper() in ("Y", "N")
               for c in CASES for a in AGENTS):
            gt = g

    pooled_rows = []
    pooled_units = []
    for case in CASES:
        rows = []
        for agent in AGENTS:
            labs = [anns[i][case].get(agent, "") for i in range(len(anns))]
            if any(v not in ("Y", "N") for v in labs):
                miss = [names[i] for i, v in enumerate(labs) if v not in ("Y", "N")]
                print(f"[warn] {case}/{agent}: missing from {miss} — item dropped")
                continue
            rows.append(labs)
            pooled_units.append(labs)
        pooled_rows += rows
        if not rows:
            print(f"\n== {case} ==  no complete items")
            continue
        oa = observed_agreement(rows)
        fk = fleiss_kappa(rows) if len({len(r) for r in rows}) == 1 else float("nan")
        ka = krippendorff_alpha_nominal(rows)
        print(f"\n== {case} ==  items={len(rows)} raters={len(anns)}  "
              f"observed={oa:.3f}  Fleiss_kappa={fk:.3f}  Krippendorff_alpha={ka:.3f}")

    oa = observed_agreement(pooled_rows)
    fk = fleiss_kappa(pooled_rows) if len({len(r) for r in pooled_rows}) == 1 else float("nan")
    ka = krippendorff_alpha_nominal(pooled_units)
    print(f"\n== POOLED (all external cases) ==  items={len(pooled_rows)} "
          f"raters={len(anns)}  observed={oa:.3f}  Fleiss_kappa={fk:.3f}  "
          f"Krippendorff_alpha={ka:.3f}")

    if gt:
        print("\n--- vs author ground truth (Cohen's kappa) ---")
        for i in range(len(anns)):
            a, b = [], []
            for case in CASES:
                for agent in AGENTS:
                    ga = str(gt[case].get(agent, "")).strip().upper()
                    aa = anns[i][case].get(agent, "")
                    if ga in ("Y", "N") and aa in ("Y", "N"):
                        a.append(ga)
                        b.append(aa)
            print(f"  {names[i]}: kappa_vs_author={cohen_kappa(a, b):.3f}  n={len(a)}")
    else:
        print("\n(author_ground_truth.json not filled — skipping annotator-vs-author. "
              "Run with no args to create the template.)")


if __name__ == "__main__":
    main()
